from __future__ import annotations
from pathlib import Path
from typing import overload, Sequence, Optional, List, Union
from openpyxl import load_workbook

@overload
def get_name_3gpp(
    folder_path: Union[str, Path],
    filename: str,
    line_numbers: int,
    *,
    sheet_name: Union[str, int, None] = None,
) -> Optional[str]: ...
@overload
def get_name_3gpp(
    folder_path: Union[str, Path],
    filename: str,
    line_numbers: Sequence[int],
    *,
    sheet_name: Union[str, int, None] = None,
) -> List[Optional[str]]: ...

def get_name_3gpp(
    folder_path: Union[str, Path],
    filename: str,
    line_numbers: Union[int, Sequence[int]],
    *,
    sheet_name: Union[str, int, None] = None,
):
    """
    高速版: .xlsx の A列（1列目）から、与えられた行番号(1始まり)の値を取得。
    - read_only=True & values_only=True でストリーム走査し、最小行〜最大行を1回だけ読む。
    - 返り値は strip() 済み（前後の空白や改行を除去）。
    - 行番号 <= 0 は None。存在しない行も None。
    - 複数指定時は入力順を維持。
    """
    path = Path(folder_path) / filename
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f".xlsx のみ対応です: {path.name}")
    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {path}")

    # 単数/複数の吸収
    single_input = isinstance(line_numbers, int)
    targets: List[int] = [int(line_numbers)] if single_input else [int(n) for n in line_numbers]

    # 0以下は即None、正の行のみ処理対象
    pos_targets = [n for n in targets if n > 0]
    if not pos_targets:
        # すべて <=0 の場合は全部 None を返す
        out = [None for _ in targets]
        return out[0] if single_input else out

    # 入力位置を維持するための結果バッファとインデックス
    result_map: dict[int, Optional[str]] = {n: None for n in pos_targets}

    # 走査範囲（1回のストリームで終了させる）
    min_row = min(pos_targets)
    max_row = max(pos_targets)
    target_set = set(pos_targets)         # 存在チェック O(1)
    remaining = set(pos_targets)          # 未充足行

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            ws = wb.active
        elif isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        # A列のみ、min_row..max_row を一度だけストリーム
        # NOTE: read_only ではランダムアクセス不可のため、1パスにまとめるのが最速
        for offset, (val,) in enumerate(
            ws.iter_rows(min_row=min_row, max_row=max_row, min_col=5, max_col=5, values_only=True),
            start=min_row
        ):
            if offset in target_set:
                result_map[offset] = (None if val is None else str(val).strip())
                remaining.discard(offset)
                if not remaining:  # すべて拾えたら即終了
                    break
    finally:
        wb.close()

    # 入力順に整形（<=0 は None のまま）
    out: List[Optional[str]] = [
        (result_map.get(n) if n > 0 else None) for n in targets
    ]
    return out[0] if single_input else out

if __name__ == "__main__":
    folder_abs_path = r"C:\Users\yohei\Downloads\特願20XX-XXXXXX_20251018\EXCEL"
    filename = "3gpp_doc_list.xlsx"
    rows = get_name_3gpp(folder_abs_path, filename,[2,5])
    # 結果を表示（例: 2,4,6）
    print(rows)