from __future__ import annotations
from pathlib import Path
from typing import Optional, Any
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook

def make_doclist_from_3gpp_excel(
    filepath_in: str | Path,
    filename_in: str,
    filepath_out: str | Path,
    filename_out: str,
    sheet_name: Optional[str] = None,
) -> Path:
    in_path  = Path(filepath_in)  / filename_in
    out_path = Path(filepath_out) / filename_out
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not in_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {in_path}")

    # ---- utils --------------------------------------------------------------
    def clean(s: Any) -> str:
        if s is None:
            return ""
        s = str(s).replace("\u00a0", " ")
        s = re.sub(r"[\r\n\t]+", " ", s)
        s = re.sub(r"\s{2,}", " ", s)
        return s.strip()

    def fmt_date(v: Any) -> str:
        if isinstance(v, datetime):
            # 00:00 の場合は日付のみ
            if v.hour == 0 and v.minute == 0 and v.second == 0:
                return v.strftime("%Y/%m/%d")
            return v.strftime("%Y/%m/%d %H:%M")
        # 文字列は軽く整形して返す
        s = clean(v)
        return s

    # ---- openpyxl: 直読み（hyperlink取得のため read_only=False が重要） ----
    wb = load_workbook(in_path, data_only=True, read_only=False)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    # ヘッダ検出（上から20行までで発見）
    header_row = None
    name2col: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=20):
        labels = [clean(c.value).lower() for c in row]
        if not any(labels):
            continue
        tmp = {lbl: i + 1 for i, lbl in enumerate(labels) if lbl}
        if {"tdoc", "title", "source"} & set(tmp.keys()):  # 必須の一部が揃えばOK
            header_row = row[0].row
            name2col = tmp
            break
    if header_row is None:
        raise ValueError("ヘッダ行（TDoc/Title/Source 等）が見つかりません。")

    def need(key_lower: str) -> int:
        if key_lower not in name2col:
            raise ValueError(f"必須列が見つかりません: {key_lower}")
        return name2col[key_lower]

    c_tdoc   = need("tdoc")
    c_title  = need("title")
    c_source = need("source")
    c_agenda = need("agenda item")  # 必須
    c_uploaded = name2col.get("uploaded")
    c_reserve  = name2col.get("reservation date")
    if c_uploaded is None and c_reserve is None:
        raise ValueError("日付列が見つかりません（Uploaded または Reservation date）")

    # ---- 出力ブック準備 ----
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Sheet1"
    out_ws.append(["URL", "著者", "タイトル", "日付", "文献番号", "アジェンダアイテム"])

    # ---- 本体走査（必要な列だけアクセス） ----
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        cell_tdoc = row[c_tdoc - 1]
        doc_no = clean(cell_tdoc.value)
        if not doc_no:  # 空行はスキップ
            continue

        # URL: hyperlink → 無ければ値からURLを拾う
        url = ""
        link = getattr(cell_tdoc, "hyperlink", None)
        if link:
            url = clean(getattr(link, "target", link))
        if not url:
            m = re.search(r"https?://\S+", doc_no)
            if m:
                url = m.group(0).rstrip(')"\'>]')

        title  = clean(row[c_title  - 1].value) if c_title  else ""
        author = clean(row[c_source - 1].value) if c_source else ""
        agenda = clean(row[c_agenda - 1].value) if c_agenda else ""

        # 日付: Uploaded 優先 → 無ければ Reservation date
        raw_date = None
        if c_uploaded:
            raw_date = row[c_uploaded - 1].value
        if (raw_date is None or raw_date == "") and c_reserve:
            raw_date = row[c_reserve - 1].value
        date_str = fmt_date(raw_date)

        out_ws.append([url, author, title, date_str, doc_no, agenda])

    out_wb.save(out_path)
    wb.close()
    return out_path
