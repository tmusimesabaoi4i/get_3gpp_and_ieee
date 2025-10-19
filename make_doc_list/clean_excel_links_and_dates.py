from __future__ import annotations
from pathlib import Path
import re
from datetime import datetime
from typing import Optional, Any
from openpyxl import load_workbook
from openpyxl.styles import Font

# 英語3文字月（必要なら 'SEPT':9 などを追加可）
_MON = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT":10, "NOV":11, "DEC":12
}

def _clean(s: Any) -> str:
    if s is None: return ""
    s = str(s).replace("\u00a0", " ")
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"\s{2,}", " ", s)
    return s.strip()

def _fmt_yyyy_mm_dd(dt: datetime) -> str:
    return f"{dt.year:04d}/{dt.month:02d}/{dt.day:02d}"

def _parse_date_like(value: Any, text_hint: Optional[str] = None) -> Optional[str]:
    """
    次のいずれも 'YYYY/MM/DD' 文字列で返す（時刻・タイムゾーンは無視）:
      - datetime
      - 'YYYY/M/D'（任意で ' HH:MM[:SS]' と ' TZ'）
      - 'DD-Mon-YY' / 'DD-Mon-YYYY'（任意で ' HH:MM[:SS]' と ' TZ'）
        例: '10-Sep-24', '10-Sep-2024 16:27:37 ET'
    """
    # 1) datetime 型
    if isinstance(value, datetime):
        return _fmt_yyyy_mm_dd(value)

    s = _clean(value if value is not None else text_hint)
    if not s:
        return None

    # 2) YYYY/M/D [HH:MM[:SS]] [TZ]
    m = re.match(
        r"^\s*(\d{4})/(\d{1,2})/(\d{1,2})"                # 年/月/日
        r"(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?"               # 任意の時刻
        r"(?:\s+[A-Za-z]{1,5})?\s*$",                     # 任意のタイムゾーン
        s
    )
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return _fmt_yyyy_mm_dd(datetime(y, mo, d))
        except ValueError:
            return None

    # 3) DD-Mon-YY(YY) [HH:MM[:SS]] [TZ]
    m = re.match(
        r"^\s*(\d{1,2})-([A-Za-z]{3})-(\d{2}|\d{4})"      # 日-月略-年(2/4桁)
        r"(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?"               # 任意の時刻
        r"(?:\s+[A-Za-z]{1,5})?\s*$",                     # 任意のタイムゾーン
        s
    )
    if m:
        d = int(m.group(1))
        mon = _MON.get(m.group(2).upper())
        year_token = m.group(3)
        if mon:
            y = int(year_token) if len(year_token) == 4 else 2000 + int(year_token)  # 24→2024
            try:
                return _fmt_yyyy_mm_dd(datetime(y, mon, d))
            except ValueError:
                return None

    return None

def clean_excel_links_and_dates(
    excel_path: str | Path,
    filename: str,
    *,
    sheet_name: Optional[str] = None,
) -> Path:
    """A列リンク除去 / D列日付を 'YYYY/MM/DD' テキストに正規化（DD,MMはゼロ詰め2桁）"""
    in_path = Path(excel_path) / filename
    if not in_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {in_path}")

    wb = load_workbook(in_path, read_only=False, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    max_row = ws.max_row or 1

    # A列: ハイパーリンク削除＆テキスト化
    for col in ws.iter_cols(min_col=1, max_col=1, min_row=1, max_row=max_row):
        for cell in col:
            shown = _clean(cell.value)
            if isinstance(cell.value, str) and cell.value.lstrip().upper().startswith("=HYPERLINK"):
                m = re.search(r",\s*\"(.*?)\"\s*\)\s*$", cell.value)
                if m: shown = _clean(m.group(1))
            try: cell.hyperlink = None
            except Exception: pass
            cell.value = shown
            cell.number_format = "@"
            cell.style = "Normal"
            cell.font = Font(underline=None)

    # D列: 'YYYY/MM/DD' に正規化（'10-Sep-2024 16:27:37 ET' なども対応）
    for col in ws.iter_cols(min_col=4, max_col=4, min_row=1, max_row=max_row):
        for cell in col:
            normalized = _parse_date_like(cell.value)
            if normalized is None:
                normalized = _parse_date_like(None, text_hint=_clean(cell.value))
            if normalized is not None:
                cell.number_format = "@"
                cell.value = normalized

    wb.save(in_path)
    wb.close()
    return in_path
